# -*- coding: utf-8 -*-
from trac_etl.database_func import *
import trac_etl.globals as globals

import os
import re
import getpass
from xlsx2csv import Xlsx2csv
import openpyxl
import time

def main_func():
    #########################################
    globals.dest_db_name = ""
    globals.dest_db_user = "postgres"
    globals.dest_db_host = "localhost"
    globals.dest_db_pass = ""
    globals.xlsx_path = "/home/gauss"
    #########################################

    n_steps = 5
    print("\nWelcome to the LALA UACh's TrAC ETL tool")
    print("========================================")

    assum = """
Prerequisites and definitions
-----------------------------
- Linux or MacOS operating system.
- Python >=3.5
- Installed Python packages:
  os, re, xlsx2csv, configparser, psycopg2, getpass, shutil, openpyxl.
- The input files have .xlsx extension. They are all in the same directory and are named:
  malla*.xlsx, *cursadas*.xlsx, *situac*.xlsx, *inscritas*.xlsx
  (Lowercase, camelcase, and uppercase for the file names are supported).
- The names of the *cursadas.xls* files are lexicographically sortable according with their date ranges.
- The destination database exists and it is empty (no tables, no objects).
    """
    print(assum)

    if True:

        globals.dest_db_host = input("> Enter destination database host (e.g. localhost): ")
        globals.dest_db_name = input("> Enter destination database name (e.g. lalauach_2022): ")
        globals.dest_db_user = input("> Enter user for the destination database (e.g. postgres): ")
        globals.dest_db_pass = getpass.getpass(prompt="> Enter user's password: ")

        start_time = time.time()

        if not test_connect():
            print("Can't connect to the database. Call this application again using the right credentials.")
            return

        ### STEP 1 ###
        print("\n* Step 1/"+str(n_steps)+": Creating database objects into "+globals.dest_db_name)
        create_db_main_objects()
        create_db_aux_objects()
        print(" Done")

        ### STEP 2 ###
        pre_entry_time = time.time()
        globals.xlsx_path = input("\n> Enter absolute path to the xlsx input files directory (e.g. /home/lalauach/xlsx/2022): ")
        entry_time = pre_entry_time - time.time()

        step_time = time.time()

        for file in os.listdir(globals.xlsx_path):
            os.rename(os.path.join(globals.xlsx_path, file), os.path.join(globals.xlsx_path, re.sub('[^a-zA-Z0-9 \n\._-]', '', file.replace(" ", "_"))))

        if True:
            print("\n* Step 2/"+str(n_steps)+": Excel files conversion")
            lst = os.listdir(globals.xlsx_path)
            lst.sort()
            for file in lst:
                if (file.endswith(".xlsx") or file.endswith(".XLSX")) and not file.startswith(".") :

                    if file.lower().find("malla") >= 0:
                        print("Pre-processing " + file + " file...")
                        try:
                            wb_obj = openpyxl.load_workbook(os.path.join(globals.xlsx_path, file))
                            sheet_obj = wb_obj.active
                            #sheet_obj.delete_cols(16, 18)
                            max_row=sheet_obj.max_row
                            for i in range(2, max_row):
                                cell_obj = sheet_obj.cell(row=i,column=3)
                                cell_obj.value = re.sub(' +', ' ', cell_obj.value)
                                cell_obj.value = cell_obj.value.replace("\n", " ")
                            wb_obj.save(os.path.join(globals.xlsx_path, file))
                            wb_obj.close()
                        except Exception as e:
                            print(e)
                            print ("Error : The file was not found")

                        if False:
                            with open(os.path.join(globals.xlsx_path, file) + ".csv", 'r') as f:
                                lines = f.readlines()
                            #lines = [line.replace(' ', '') for line in lines]
                            lines = [re.sub(' +', ' ', line) for line in lines]
                            f.close()

                            with open(os.path.join(globals.xlsx_path, file) + ".csv", 'w') as f:
                                f.writelines(lines)
                            f.close()

                            f = open(os.path.join(globals.xlsx_path, file) + ".csv", "rt")
                            file_content = f.read()
                            file_content = file_content.replace(chr(177)+"", "ñ") # ±
                            f.close()
                            f = open(os.path.join(globals.xlsx_path, file) + ".csv", "wt")
                            f.write(file_content)
                            f.close()

                        print(" Done")

                    elif False and file.lower().find("situac") >= 0:
                        print("Pre-processing " + file + " file...")
                        try:
                            wb_obj = openpyxl.load_workbook(os.path.join(globals.xlsx_path, file))
                            sheet_obj = wb_obj.active
                            sheet_obj.delete_cols(14, 16)
                            wb_obj.save(os.path.join(globals.xlsx_path, file))
                            wb_obj.close()
                        except Exception as e:
                            print(e)
                            print ("Error : The file was not found")
                        print(" Done")

                    print("Converting " + file + " to UTF-8 CSV...")
                    Xlsx2csv(os.path.join(globals.xlsx_path, file), outputencoding="utf-8").convert(os.path.join(globals.xlsx_path, file) + ".csv")

                    if False and (file.lower().find("malla") >= 0 or file.lower().find("situac") >= 0):
                        with open(os.path.join(globals.xlsx_path, file) + ".csv", 'r') as f:
                            lines = f.readlines()
                        f.close()

                        for i, element in enumerate(lines):
                            if i > 0:
                                lines[i] = str(i) + "," + lines[i]

                        with open(os.path.join(globals.xlsx_path, file) + ".csv", 'w') as f:
                            f.writelines(lines)
                        f.close()

                    print(" Done")
            print("[All files converted to CSV in", round((time.time() - step_time)/60.0, 1) ,"minutes]")

    if True:
        ### STEP 3 ###
        step_time = time.time()
        print("\n* Step 3/"+str(n_steps)+": CSVs data insertion")
        truncate_xcursadas = True
        lst = os.listdir(globals.xlsx_path)
        lst.sort()
        for file in lst:
            if file.endswith(".csv") and not file.startswith(".") :
                if file.lower().find("malla") >= 0:
                    print("Inserting xmalla data...")
                    csv_to_xmalla(os.path.join(globals.xlsx_path, file))
                    print(" Done")
                elif file.lower().find("situac") >= 0:
                    print("Inserting xsituacion_semestral data...")
                    csv_to_xsituac_sem(os.path.join(globals.xlsx_path, file))
                    print(" Done")
                elif file.lower().find("cursadas") >= 0:
                    print("Inserting xcursadas data (" + file + ")...")
                    csv_to_xcursadas(os.path.join(globals.xlsx_path, file), truncate_xcursadas)
                    if truncate_xcursadas:
                        truncate_xcursadas = False
                    print(" Done")
                elif file.lower().find("inscritas") >= 0:
                    print("Inserting inscritas2020 data (" + file + ")...")
                    csv_to_inscritas2020(os.path.join(globals.xlsx_path, file))
                    print(" Done")
        print("[CSVs inserted in database in", round((time.time() - step_time)/60.0, 1) ,"minutes]")

    ### STEP 4 ###
    step_time = time.time()
    print("\n* Step 4/"+str(n_steps)+": LALA UACh data model population")
    print("Inserting program_structure data...")
    ins_program_structure()
    print(" Done")
    print("Inserting course data...")
    ins_course()
    print(" Done")
    print("Inserting student_term data...")
    ins_student_term()
    print(" Done")
    print("Encoding student_term comments...")
    upd_student_term_comm()
    print(" Done")
    print("Inserting student_course data...")
    ins_student_course()
    print(" Done")
    print("Inserting inscritas2020 data into student_course...")
    ins_inscritas2020_into_stud_cou()
    print(" Done")
    print("Inserting inscritas2020 data into student_term...")
    ins_inscritas2020_into_stud_term()
    print(" Done")
    print("Updating student_term 2020 start_year...")
    upd_student_term_year()
    print(" Done")

    # This is accomplished in ins_student_course()
    #print("Rounding student_course grades...")
    #upd_student_course_grade()
    #print(" Done")
    print("Creating xstudent_program temporary table...")
    cre_ins_xstudent_program()
    print(" Done")
    print("Creating xtemp_sc temporary table...")
    cre_ins_xtemp_sc()
    print(" Done")
    print("Inserting student_program data...")
    ins_student_program()
    print(" Done")
    print("Inserting student data...")
    ins_student()
    print(" Done")
    print("Creating x_last_gpas temporary table...")
    cre_ins_x_last_gpas()
    print(" Done")
    print("Computing and inserting student_cluster...")
    ins_student_cluster()
    print(" Done")
    print("[Model data inserted in", round((time.time() - step_time)/60.0, 1) ,"minutes]")

    ### STEP 5 ###
    step_time = time.time()
    print("\n* Step 5/"+str(n_steps)+": Computation of courses statistics")
    print("Inserting stats by parallel groups...")
    ins_course_stats_parallel_group(truncate=True)
    print(" Done")
    print("Inserting stats by course...")
    ins_course_stats_course()
    print(" Done")
    print("Inserting stats by semester...")
    ins_course_stats_semester()
    print(" Done")
    print("Statistics computed in", round((time.time() - step_time)/60.0, 1) ,"minutes")

    print("\n========================================")
    print("LALA UACh's TrAC ETL tool completed.")
    print("[Total execution time:", round((time.time() - start_time - entry_time)/60.0, 1) ,"minutes]")
    print("Have a nice day.")
